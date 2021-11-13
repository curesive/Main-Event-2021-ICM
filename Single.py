#! python3
#Monte Carlo Sim operating on a single tournament
#Using Sheet1 from TournamentData.xlsx
#Output goes to updatedTournamentData.xlsx

from random import *
import openpyxl, pprint, sys, math, timeit
from openpyxl.styles import Font
import matplotlib.pyplot as plt
import numpy as np

def trial(weights, payouts, tournamentNum):           
        #scores is a list, the items inside the list are tuples. Each tuple contains the random score
        #and it's corresponding stack number. [(0.15533, 8), (0.18223, 3) ... ]
        #low weight = best chance of winning
        scores = sorted((random() ** weight, i) for i, weight in enumerate(weights))
        #populate finish distribution
        #weights[] is sorted from lowest number to highest, corresponds to highest stack to lowest stack
        for i in range(0, len(scores)):
                finishDistribution[tournamentNum]['Player ' + str(scores[i][1] + 1)]['Position ' + str(len(scores) - i)] += 1
                payoutList[tournamentNum]['Player ' + str(scores[i][1] + 1)].append(payouts[i])
                globalCDFList.append([])
        results = [0] * len(payouts)
        for payout, score in zip(payouts, scores): results[score[1]] = payout
        return results

def sicm(wbName, tournamentIndex):
        wb = openpyxl.load_workbook(wbName)
        sheet = wb['Sheet1']

        global globalPayouts, globalStacks
        globalPayouts.append([])
        globalStacks.append([])

        for i in range(0, 399):
                globalPayouts[tournamentIndex].append(sheet['D' + str(i + 4)].value)
                globalStacks[tournamentIndex].append(sheet['B' + str(i + 4)].value)
        globalPayouts[tournamentIndex] = list(filter(None, globalPayouts[tournamentIndex]))
        globalPayouts[tournamentIndex].sort() #arranges values in payouts from lowest number to highest
        globalStacks[tournamentIndex] = list(filter(None, globalStacks[tournamentIndex]))

        totalChips = sum(globalStacks[tournamentIndex])
        trimPayouts(tournamentIndex)
        payoutList.append({})
        finishDistribution.append({})
        finishProbability.append({})
        globalVarianceList.append({})

        avg = sum(globalStacks[tournamentIndex]) / float(len(globalStacks[tournamentIndex])) #avg is average stack size
        weights = [avg / s for s in globalStacks[tournamentIndex]] #list with items equal to the average stack divided by each individual stack

        for i in range(1, (len(weights) + 1)):
                finishDistribution[tournamentIndex].setdefault('Player ' + str(i), {})
                finishProbability[tournamentIndex].setdefault('Player ' + str(i), {})
                globalVarianceList[tournamentIndex].setdefault('Player ' + str(i), {})
                globalVarianceList[tournamentIndex]['Player ' + str(i)].setdefault('Variance')
                globalVarianceList[tournamentIndex]['Player ' + str(i)].setdefault('Percentage of Chips', (globalStacks[tournamentIndex][i - 1] / totalChips))
                globalVarianceList[tournamentIndex]['Player ' + str(i)].setdefault('Standard Deviation')
                globalVarianceList[tournamentIndex]['Player ' + str(i)].setdefault('EV')                
                payoutList[tournamentIndex].setdefault('Player ' + str(i), [])
                for j in range(1, (len(globalStacks[tournamentIndex]) + 1)):
                        finishDistribution[tournamentIndex]['Player ' + str(i)].setdefault('Position ' + str(j), 0)
                        finishProbability[tournamentIndex]['Player ' + str(i)].setdefault('Position ' + str(j), 0.0)
        return [sum(player) / float(globalTrials) for player in zip(*( 
                trial(weights, globalPayouts[tournamentIndex], 0) for i in range(globalTrials)
        ))]

def printDist(wbName, actionPercentage, markup, tournamentIndex):
        #need to update MEMS to work with integer finDist
        wb = openpyxl.load_workbook(wbName)
        sheet = wb['Sheet' + str(globalOutputSheetNumber)]
        rowNum = 1
        numPayouts = len(globalPayouts[tournamentIndex])        
        
        for playerNum, player in enumerate(finishDistribution[tournamentIndex]):
                sheet.cell(row = rowNum, column = 6).value = player
                sheet.cell(row = rowNum, column = 7).value = 'Finish %'
                sheet.cell(row = rowNum, column = 8).value = 'Normalized Result'
                sheet.cell(row = playerNum + 4, column = 2).value = globalStacks[tournamentIndex][playerNum]             
                sheet.cell(row = playerNum + 4, column = 3).value = icmValues[tournamentIndex][playerNum]
                sheet.cell(row = playerNum + 4, column = 5).value = adjSValueList[playerNum]
                for finishNum, finish in enumerate(finishDistribution[tournamentIndex][player]):
                        sheet.cell(row = (rowNum + finishNum + 1), column = 6).value = finishNum + 1
                        #changed this to print integer value of finishDistribution            
                        #sheet.cell(row = (rowNum + finishNum + 1), column = 7).value = (finishDistribution[tournamentIndex][player][finish] / globalTrials * 100)
                        sheet.cell(row = (rowNum + finishNum + 1), column = 7).value = (finishDistribution[tournamentIndex][player][finish])
                        sheet.cell(row = (rowNum + finishNum + 1), column = 8).value = ((globalPayouts[tournamentIndex][numPayouts - finishNum - 1] * actionPercentage) - icmValues[tournamentIndex][playerNum] * actionPercentage * markup) / (icmValues[tournamentIndex][playerNum] * actionPercentage * markup)
                rowNum += (len(finishDistribution[tournamentIndex][player]) + 2)
        wb.save('updatedTournamentData.xlsx')
        wb.save(wbName)
        return

def rebuildPayoutList(tournamentIndex):        
        for i in range(1, (len(payoutList[tournamentIndex]) + 1)):
                counter = 0
                actionPlayer = 'Player ' + str(i)
                for j in range(1, (len(finishDistribution[tournamentIndex][actionPlayer]) + 1)):
                        for k in range(0, finishDistribution[tournamentIndex][actionPlayer]['Position ' + str(j)]):
                                if((k + counter) >= len(payoutList[tournamentIndex][actionPlayer])):
                                        payoutList[tournamentIndex][actionPlayer].append(globalPayouts[tournamentIndex][len(globalPayouts[tournamentIndex]) - j])
                                else:
                                        payoutList[tournamentIndex][actionPlayer][k + counter] = globalPayouts[tournamentIndex][len(globalPayouts[tournamentIndex]) - j]
                        counter += finishDistribution[tournamentIndex][actionPlayer]['Position ' + str(j)]
                shuffle(payoutList[tournamentIndex][actionPlayer])
        return

def buildActionProfitList(actionPlayer, actionPrice):
        actionProfitList = []
        actionProfitListNormalized = []        
        actionResultList = []
        actionResultList.append(payoutList[0][actionPlayer][0] * globalActionPercentage)
        actionProfitList.append(actionResultList[0] - actionPrice)
        actionProfitListNormalized.append(actionProfitList[0] / actionPrice)

        for i in range(1, len(payoutList[0][actionPlayer])):
                actionResultList.append(payoutList[0][actionPlayer][i] * globalActionPercentage)
                actionProfitList.append((actionResultList[i] - actionPrice) + actionProfitList[i - 1]) 
                actionProfitListNormalized.append(actionProfitList[i] / actionPrice) 
        return actionResultList, actionProfitList, actionProfitListNormalized

def printActionSim(actionPlayer, actionPrice, actionPercentage, sampleRunCount, trialsPerSampleRun):
        wb = openpyxl.load_workbook('updatedTournamentData.xlsx')
        sheet = wb['Sheet1']       
        
        trialProfitDictionary = {}
        trialProfitDictionaryNormalized = {}         
        actionResultList, actionProfitList, actionProfitListNormalized = buildActionProfitList(actionPlayer, actionPrice)

        for k in range(0, len(actionProfitList)):
                sheet.cell(row = k + 2, column = 25).value = actionProfitList[k]
                sheet.cell(row = k + 2, column = 26).value = actionProfitListNormalized[k]
                
        counter = 0 #counter keeps track of where we are in actionResultList as we slice it up
        for i in range(0, sampleRunCount):
                trialProfitDictionary.setdefault(i, [])
                trialProfitDictionaryNormalized.setdefault(i, [])
                for j in range(0, trialsPerSampleRun):
                        if(j == 0):
                               trialProfitDictionary[i].append(actionResultList[counter] - actionPrice)
                               trialProfitDictionaryNormalized[i].append((trialProfitDictionary[i][j]) / actionPrice)
                               counter += 1
                        else:                                                
                                trialProfitDictionary[i].append((actionResultList[counter] - actionPrice) + trialProfitDictionary[i][j - 1])
                                trialProfitDictionaryNormalized[i].append((trialProfitDictionary[i][j] / actionPrice))
                                counter += 1
        colCounter = 0
        for i in range(0, len(trialProfitDictionary)):
                sheet.cell(row = 1, column = 28 + colCounter).value = ('Trial #' + str(i + 1) + ':')
                sheet.cell(row = 1, column = 29 + colCounter).value = 'Sum $ cashed:'
                sheet.cell(row = 1, column = 30 + colCounter).value = 'Normalized $ cashed:'
                for j in range(0, len(trialProfitDictionary[i])):
                        sheet.cell(row = j + 2, column = 29 + colCounter).value = trialProfitDictionary[i][j]
                        sheet.cell(row = j + 2, column = 30 + colCounter).value = trialProfitDictionaryNormalized[i][j]
                colCounter += 4                
        colCount = 29 #starting column that holds data for trial #1 of variance trials
        seriesCount = 1
        refObj = []
        refObjNormalized = []
        seriesObj = []
        seriesObjNormalized = []
        chartObj = openpyxl.chart.LineChart()
        chartObj.title = ('Winnings in $ over ' + str(sampleRunCount) + ' trials of ' + str(trialsPerSampleRun) + ' tournaments')
        chartObjNormalized = openpyxl.chart.LineChart()
        chartObjNormalized.title = ('Winnings in buyins over ' + str(sampleRunCount) + ' trials of ' + str(trialsPerSampleRun) + ' tournaments')        
        for i in range(0, sampleRunCount):                
                refObj.append(openpyxl.chart.Reference(sheet, min_col = colCount, min_row = 2, max_col = colCount, max_row = (trialsPerSampleRun + 1)))
                refObjNormalized.append(openpyxl.chart.Reference(sheet, min_col = colCount + 1, min_row = 2, max_col = colCount + 1, max_row = trialsPerSampleRun + 1))
                seriesObj.append(openpyxl.chart.Series(refObj[i], title = 'Trial ' + str(seriesCount)))
                seriesObjNormalized.append(openpyxl.chart.Series(refObjNormalized[i], title = 'Trial ' + str(seriesCount)))
                chartObj.append(seriesObj[i])
                chartObjNormalized.append(seriesObjNormalized[i])
                colCount = colCount + 4
                seriesCount += 1               
        sheet.add_chart(chartObj, 'O2')
        sheet.add_chart(chartObjNormalized, 'O36')                
        wb.save('updatedTournamentData.xlsx')
        print('Price paid for action is ' + str(actionPrice))
        return

def buildProbabilityMassFunction(betsPerSlice, playerIndex, tournamentIndex):
        actionPlayer = 'Player ' + str(playerIndex)        
        actionPrice = icmValues[tournamentIndex][playerIndex - 1] * globalActionPercentage * globalActionMarkup        
        actionResultList, actionProfitList, actionProfitListNormalized = buildActionProfitList(actionPlayer, actionPrice)        
        numSlices = 200000
        terminalBankrollSizeList, roundedTerminalBankrollSizeList = calcPMF(numSlices, betsPerSlice, actionProfitListNormalized)
        countList = []
        xMin = -1 * globalBRSize
        xMax = 46 #one over the maximum we want to use        
        for i in range(xMin, xMax): #builds countlist, which is the PMF
                countList.append(roundedTerminalBankrollSizeList.count(i) / numSlices) #normalized to probability        
        #builds CDF
        globalCDFList[playerIndex - 1].append(countList[0])
        for i in range(1, (xMax - xMin)):
                globalCDFList[playerIndex - 1].append(globalCDFList[playerIndex - 1][i - 1] + countList[i])
        pmfArray = np.array(countList)
        if(doPlotPMF == 'y'):
                plotPMF(xMin, xMax, countList, playerIndex, betsPerSlice)        
        return pmfArray, xMin, xMax

def printVarianceSingleTournament():
        wb = openpyxl.load_workbook('updatedTournamentData.xlsx')
        sheet = wb['Sheet1']
        boldStandardFont = Font(size = 11, bold = True)
        sheet.cell(row = 50, column = 10).value = 'single action buy'
        sheet['J50'].font = boldStandardFont
        sheet.cell(row = 50, column = 11).value = '% of Chips:'
        sheet.cell(row = 50, column = 12).value = 'Variance:'
        sheet.cell(row = 50, column = 13).value = 'Standard Deviation'
        sheet.cell(row = 50, column = 14).value = 'Normalized EV'

        rowNum = 50
        for i in range(1, len(globalVarianceList[0]) + 1):                
                sheet.cell(row = rowNum + i, column = 10).value = 'Player ' + str(i)
                sheet.cell(row = rowNum + i, column = 11).value = globalVarianceList[0]['Player ' + str(i)]['Percentage of Chips']
                sheet.cell(row = rowNum + i, column = 12).value = globalVarianceList[0]['Player ' + str(i)]['Variance']
                sheet.cell(row = rowNum + i, column = 13).value = globalVarianceList[0]['Player ' + str(i)]['Standard Deviation']
                sheet.cell(row = rowNum + i, column = 14).value = globalVarianceList[0]['Player ' + str(i)]['EV']

        wb.save('updatedTournamentData.xlsx')
        return

def printCDF(tournamentIndex):
        wb = openpyxl.load_workbook('updatedTournamentData.xlsx')
        sheet = wb['Sheet1']
        boldStandardFont = Font(size = 11, bold = True)
        rowNum = 54 + len(globalVarianceList[tournamentIndex])
        
        sheet.cell(row = rowNum, column = 10).value = str(globalTrialsPerSampleRun) + ' action buys'
        sheet['J' + str(rowNum)].font = boldStandardFont
        sheet.cell(row = rowNum, column = 12).value = 'Variance:'
        sheet.cell(row = rowNum, column = 13).value = 'Standard Deviation'
        sheet.cell(row = rowNum, column = 14).value = 'Normalized EV'

        for i in range(1, len(globalVarianceList[tournamentIndex]) + 1):
                #builds PMF and CDF to print
                pmfArray, xMin, xMax = buildProbabilityMassFunction(globalTrialsPerSampleRun, i, tournamentIndex)
                sheet.cell(row = rowNum + i, column = 10).value = 'Player ' + str(i)
                sheet.cell(row = rowNum + i, column = 11).value = globalVarianceList[tournamentIndex]['Player ' + str(i)]['Percentage of Chips']
                sheet.cell(row = rowNum + i, column = 12).value = globalVarianceList[tournamentIndex]['Player ' + str(i)]['Variance'] * globalTrialsPerSampleRun
                sheet.cell(row = rowNum + i, column = 13).value = math.sqrt(globalVarianceList[tournamentIndex]['Player ' + str(i)]['Variance'] * globalTrialsPerSampleRun)
                sheet.cell(row = rowNum + i, column = 14).value = globalVarianceList[tournamentIndex]['Player ' + str(i)]['EV'] * globalTrialsPerSampleRun

        #print CDF for each possible buyin level
        rowNum = rowNum + 4 + len(globalVarianceList[tournamentIndex])
        sheet.cell(row = rowNum, column = 10).value = str(globalTrialsPerSampleRun)+ ' action buys'

        for i in range(1, len(globalVarianceList[tournamentIndex]) + 1):                
                sheet.cell(row = rowNum, column = 10 + i).value = 'Player ' + str(i)
                for j in range(0, (-1 * xMin) + 1):
                        #print CDF data for this player index
                        sheet.cell(row = rowNum + 1 + j, column = 10 + i).value = globalCDFList[i - 1][j]                
                for k in range((-1 * xMin),(xMax - xMin)):
                        sheet.cell(row = rowNum + 2 + k, column = 10 + i).value = globalCDFList[i - 1][xMax - xMin - 1] - globalCDFList[i - 1][k]
                
        for j in range(0, (-1 * xMin) + 1):
                sheet.cell(row = rowNum + j + 1, column = 10).value = 'Lose ' + str((-1 * xMin) - j) + '+'                   
        rowNum = rowNum + 2 + (-1 * xMin)
        for k in range(0, xMax): #xMax is already incremented 1 over true xMax
                sheet.cell(row = rowNum + k, column = 10).value = 'Win ' + str(k) + '+'
        wb.save('updatedTournamentData.xlsx')
        return

def varianceCalc(tournamentIndex, playerIndex):
        #this function operates on updatedTournamentData, and displays variance details to Sheet1
        #playerIndex is 1 indexed, so use playerIndex - 1 for lists
        numPlayers = len(finishDistribution[tournamentIndex])
        numFinishes = len(finishDistribution[tournamentIndex]['Player 1'])
        numPayouts = len(globalPayouts[tournamentIndex])
        EV = [0] * numPlayers
        EVsquared = [0] * numPlayers
        VarianceSum = [0] * numPlayers
        Variance = [0] * numPlayers
        StdDev = [0] * numPlayers
        pi = [0] * numFinishes
        xi = [0] * numFinishes

        #this sim can calculate in BUYINS or DOLLARS
        runSimDollars = 0
        actionPriceInDollars = 200000

        for j in range(0, numPlayers):
                if(j == playerIndex - 1):
                        if(runSimDollars == 0):
                                actionPrice = icmValues[tournamentIndex][j] * globalActionMarkup * globalActionPercentage
                                percentageBought = globalActionPercentage
                        else:
                                actionPrice = actionPriceInDollars
                                percentageBought = actionPrice / (icmValues[tournamentIndex][j] * globalActionMarkup)
                else:
                        if(runSimDollars == 0):
                                actionPrice = icmValues[tournamentIndex][j] * globalActionPercentage
                                percentageBought = globalActionPercentage
                        else:
                                actionPrice = actionPriceInDollars
                                percentageBought = actionPrice / icmValues[tournamentIndex][j]

                for i in range(0, numFinishes):
                        actionResult = globalPayouts[tournamentIndex][numPayouts - 1 - i] * percentageBought
                        pi[i] = round((finishDistribution[tournamentIndex]['Player ' + str(j + 1)]['Position ' + str(i + 1)]) / globalTrials, 5)
                        if(runSimDollars == 0):
                                xi[i] = round((actionResult - actionPrice) / actionPrice, 5)
                        else:
                                xi[i] = round((actionResult - actionPrice), 5)
                        VarianceProduct = pi[i] * xi[i] * xi[i]
                        EVProduct = pi[i] * xi[i]
                        VarianceSum[j] = VarianceSum[j] + VarianceProduct
                        EV[j] = EV[j] + EVProduct
                EVsquared[j] = EV[j] * EV[j]
                Variance[j] = VarianceSum[j] - EVsquared[j]
                StdDev[j] = math.sqrt(Variance[j])
                globalVarianceList[tournamentIndex]['Player ' + str(j + 1)]['Variance'] = Variance[j]
                globalVarianceList[tournamentIndex]['Player ' + str(j + 1)]['Standard Deviation'] = StdDev[j]
                globalVarianceList[tournamentIndex]['Player ' + str(j + 1)]['EV'] = EV[j]        
        printVarianceSingleTournament()
        printCDF(tournamentIndex)
        return

def edgeTransform(edge, playerIndex, tournamentIndex):
        #ICMDifference holds the $ amount that we need to increase the players sum(payoutList) by to arrive at our target edge
        ICMDifference = icmValues[tournamentIndex][playerIndex - 1] * edge - icmValues[tournamentIndex][playerIndex - 1]
        
        if(playerIndex < len(icmValues[tournamentIndex])):                
                recipPlayerIndex = playerIndex + 1
        else:
                recipPlayerIndex = playerIndex - 1
        recipICMValue = icmValues[tournamentIndex][recipPlayerIndex - 1] - ICMDifference
        adjustmentSum = 0
        payoutLength = len(globalPayouts[tournamentIndex])

        #Split ICMDifference into chunks, signifying the amount of $ added to ICMValue corresponding to each increased finish position
        #need to adjust this for even number of players at the table
        midpoint = round(len(icmValues[tournamentIndex]) / 2)
        numSteps = midpoint - 1
        weightedIncrement = 1
        while (adjustmentSum < ICMDifference):
                #take last place, turn it into midpoint+1, 2nd to last -> midpoint + 2, etc
                for i in range(1, numSteps + 1):
                        if(i <= 2):
                                weightedIncrement = 2
                        else:
                                weightedIncrement = 1                        
                        finishDistribution[tournamentIndex]['Player ' + str(playerIndex)]['Position ' + str(i)] += weightedIncrement                                      
                        finishDistribution[tournamentIndex]['Player ' + str(playerIndex)]['Position ' + str(midpoint + i)] -= weightedIncrement 
                              
                        finishDistribution[tournamentIndex]['Player ' + str(recipPlayerIndex)]['Position ' + str(i)] -= weightedIncrement 
                        finishDistribution[tournamentIndex]['Player ' + str(recipPlayerIndex)]['Position ' + str(midpoint + i)] += weightedIncrement                        
                        adjustmentSum += ((globalPayouts[tournamentIndex][payoutLength - i] - globalPayouts[tournamentIndex][payoutLength - (midpoint + i)]) * weightedIncrement) / globalTrials
        return

def calcRiskOfRuin(playerIndex, tournamentIndex):
        globalRoRList.append({}) #adds dictionary for tournament index, just 0 placeholder
        tournamentCountStepList = []
        buyinThresholdMax = 90 #maximum win checked
        tournamentCountStep = 10 #evaluate the RoR in steps of i*tournamentCountStep. 10,20,30,40... tournaments
        tournamentCountMin = 100
        tournamentCountMax = 200
        CDFLoopCount = int((tournamentCountMax - tournamentCountMin) / tournamentCountStep)
        
        actionPlayer = 'Player ' + str(playerIndex)        
        globalRoRList[tournamentIndex].setdefault('Player ' + str(playerIndex), {})
        actionPrice = icmValues[tournamentIndex][playerIndex - 1] * globalActionPercentage * globalActionMarkup        
        actionResultList, actionProfitList, actionProfitListNormalized = buildActionProfitList(actionPlayer, actionPrice)

        CDFXList = []
        CDFXListWin = []        
        for i in range(0, (globalBRSize + 1)):
                globalRoRList[tournamentIndex]['Player ' + str(playerIndex)].setdefault('Lose ' + str(i) + '+ Buyins', [])
                CDFXList.append((-1 * globalBRSize) + i)
        for i in range(0, buyinThresholdMax + 1):
                globalRoRList[tournamentIndex]['Player ' + str(playerIndex)].setdefault('Win ' + str(i) + '+ Buyins', [])
                CDFXListWin.append(i)

        #run a single CDF, if not, executes the loop code after this block
        if(doLoopRoR == 'n'):
                tempCDFList, countList, xMin, xMax = buildCDF(globalTrialsPerSampleRun, playerIndex, actionProfitListNormalized, buyinThresholdMax, tournamentIndex)
                globalRoRList[tournamentIndex][actionPlayer].setdefault('Bets = ' + str(globalTrialsPerSampleRun), [])
                for j in range(0, globalBRSize + 1):
                        tempProbability = tempCDFList[playerIndex - 1][j]
                        globalRoRList[tournamentIndex][actionPlayer]['Lose ' + str(globalBRSize - j) + '+ Buyins'].append(tempProbability)
                        globalRoRList[tournamentIndex][actionPlayer]['Bets = ' + str(globalTrialsPerSampleRun)].append(globalRoRList[tournamentIndex][actionPlayer]['Lose ' + str(globalBRSize - j) + '+ Buyins'][0])
                for j in range(0, buyinThresholdMax + 1):
                        tempProbability = 1 - tempCDFList[playerIndex - 1][j + globalBRSize]
                        globalRoRList[tournamentIndex][actionPlayer]['Win ' + str(j) + '+ Buyins'].append(tempProbability)
                        globalRoRList[tournamentIndex][actionPlayer]['Bets = ' + str(globalTrialsPerSampleRun)].append(globalRoRList[tournamentIndex][actionPlayer]['Win ' + str(j) + '+ Buyins'][0])
                plotCDF(playerIndex, globalTrialsPerSampleRun, CDFXList, CDFXListWin)
                #def plotPMF(xMin, xMax, countList, playerIndex, betsPerSlice):
                plotPMF(xMin, xMax, countList, playerIndex, globalTrialsPerSampleRun)
                sys.exit()
        
        #for CDFLoopCount = 1, this calls buildCDF twice, not sure that is optimal. 
        for i in range(0, CDFLoopCount + 1):
                tournamentCountStepList.append(int(tournamentCountMin + (i * tournamentCountStep)))
                #build CDF for i*tournamentCountStep tournament buys
                tempCDFList, countList, xMin, xMax = buildCDF((tournamentCountMin + (i * tournamentCountStep)), playerIndex, actionProfitListNormalized, buyinThresholdMax, tournamentIndex)
                for j in range(0, (globalBRSize + 1)):
                        #starting at lose xMin+ buyins, work down to 0+ buyins
                        tempProbability = tempCDFList[playerIndex - 1][j]
                        globalRoRList[tournamentIndex][actionPlayer]['Lose ' + str(globalBRSize - j) + '+ Buyins'].append(tempProbability)
                for j in range(0, buyinThresholdMax + 1):
                        tempProbability = 1 - tempCDFList[playerIndex - 1][j + globalBRSize]
                        globalRoRList[tournamentIndex][actionPlayer]['Win ' + str(j) + '+ Buyins'].append(tempProbability)
        
        #this loop will build CDF for each player for each tournamentCountStep
        for i in range(0, CDFLoopCount + 1):
                globalRoRList[tournamentIndex][actionPlayer].setdefault('Bets = ' + str(tournamentCountMin + (tournamentCountStep * i)), [])
                for j in range(0, globalBRSize + 1):
                        #fill in lose 0+ thru lose 20+ buyins in CDF
                        globalRoRList[tournamentIndex][actionPlayer]['Bets = ' + str(tournamentCountMin + (tournamentCountStep * i))].append(globalRoRList[tournamentIndex][actionPlayer]['Lose ' + str(globalBRSize - j) + '+ Buyins'][i])
                for j in range(0, buyinThresholdMax + 1):
                        globalRoRList[tournamentIndex][actionPlayer]['Bets = ' + str(tournamentCountMin + (tournamentCountStep * i))].append(globalRoRList[tournamentIndex][actionPlayer]['Win ' + str(j) + '+ Buyins'][i])        
        plotCDF(playerIndex, (tournamentCountMin + (tournamentCountStep * i)), CDFXList, CDFXListWin)
        return

def calcPMF(numSlices, betsPerSlice, actionProfitListNormalized):
        #This function is only used for single player simulations
        #This function uses globalBRSize, which is in BUYINS
        trials = len(actionProfitListNormalized)
        terminalBankrollSizeList = []
        roundedTerminalBankrollSizeList = []
        for i in range(0, numSlices):
                startingIndex = randint(0, (trials - betsPerSlice))                
                #take a random slice of length betsPerSlice out of actionProfitListNormalized
                tempList = actionProfitListNormalized[startingIndex: startingIndex + betsPerSlice : 1]
                #check if the bankroll went to 0 at any point in the slice
                tempMin = min(tempList)
                if(tempMin - tempList[0] <= -1 * globalBRSize): #bankroll hit 0 flag
                        terminalBankrollSizeList.append(globalBRSize * -1)                      
                else: #no ruin                        
                        terminalBankrollSizeList.append(tempList[betsPerSlice - 1] - tempList[0]) #stores net buyins in this list
                roundedTerminalBankrollSizeList.append(int(round(terminalBankrollSizeList[i]))) #sorts each ending bankroll value into integer buckets
        return terminalBankrollSizeList, roundedTerminalBankrollSizeList

def buildCDF(betsPerSlice, playerIndex, actionProfitListNormalized, buyinThresholdMax, tournamentIndex):      
        #careful with playerIndex in this function. not 0 indexed
        numSlices = 200000
        CDFList = []
        xMin = -1 * globalBRSize
        xMax = buyinThresholdMax + 1 #+1 for loops
        
        for i in range(0, len(icmValues[tournamentIndex])):
                CDFList.append([])

        terminalBankrollSizeList, roundedTerminalBankrollSizeList = calcPMF(numSlices, betsPerSlice, actionProfitListNormalized)
        countList = []

        for i in range(xMin, xMax):
                countList.append(roundedTerminalBankrollSizeList.count(i) / numSlices) #normalized to probability
        CDFList[playerIndex - 1].append(countList[0])
        for i in range(1, (xMax - xMin)):
                CDFList[playerIndex - 1].append(CDFList[playerIndex - 1][i - 1] + countList[i])
        return CDFList, countList, xMin, xMax

def plotRoR(playerIndex, tournamentCountStepList, buyinThreshold):
        plt.plot(tournamentCountStepList, globalRoRList[0]['Player ' + str(playerIndex)]['Lose ' + str(buyinThreshold) + '+ Buyins'])
        plt.title('% chance that Player ' + str(playerIndex) + ' loses ' + str(buyinThreshold) + ' or more buyins over n bets')
        plt.xlabel('# bets made in succession')
        plt.ylabel('Probability')
        plt.annotate('Bankroll = ' + str(globalBRSize) + ' buyins', xy = (0.1, 0.9), xycoords = 'axes fraction')
        plt.grid(True)
        plt.show()      
        return

def plotCDF(playerIndex, numBets, CDFXList, CDFXListWin):
        plt.subplot(1, 2, 1)
        plt.plot(CDFXList, globalRoRList[0]['Player ' + str(playerIndex)]['Bets = ' + str(numBets)][0 : (len(CDFXList))])
        plt.title('% chance that buying action in Player ' + str(playerIndex) + ' loses n buyins or more over ' + str(numBets) + ' bets in succession')
        plt.xlabel('# Buyins')
        plt.ylabel('Probability')
        plt.annotate('Bankroll = ' + str(globalBRSize) + ' buyins', xy = (0.1, 0.9), xycoords = 'axes fraction')
        plt.grid(True)

        plt.subplot(1, 2, 2)
        plt.plot(CDFXListWin, globalRoRList[0]['Player ' + str(playerIndex)]['Bets = ' + str(numBets)][len(CDFXList):(len(globalRoRList[0]['Player ' + str(playerIndex)]['Bets = ' + str(numBets)]))])
        plt.title('% chance that buying action in Player ' + str(playerIndex) + ' wins n buyins or more over ' + str(numBets) + ' bets in succession')
        plt.xlabel('# Buyins')
        plt.ylabel('Probability')
        plt.annotate('Bankroll = ' + str(globalBRSize) + ' buyins', xy = (0.1, 0.9), xycoords = 'axes fraction')
        plt.grid(True)
        
        plt.show()
        return

def plotPMF(xMin, xMax, countList, playerIndex, betsPerSlice):
        pmfListX = []         
        for i in range(0, xMax - xMin):
                pmfListX.append(xMin + i)

        plt.plot(pmfListX, countList)
        plt.title('Bankroll growth for player ' + str(playerIndex) + ' over ' + str(betsPerSlice) + ' bets')
        plt.xlabel('Buyins')
        plt.ylabel('Probability')
        plt.grid(True)
        plt.show()
        return

def readICMValues(wbName, sheetName, tournamentIndex):
        #this function reads in finishDistribution and ICMValues from previously run Excel sheet, to save time on simulation
        wb = openpyxl.load_workbook(wbName)
        sheet = wb[sheetName]
        rowNum = 1
        icmRowNum = 3
        
        ICMValues = []
        global globalPayouts, globalStacks
        globalPayouts.append([])
        globalStacks.append([])

        for i in range(0, 399):
                globalPayouts[tournamentIndex].append(sheet['D' + str(i + 4)].value)
                globalStacks[tournamentIndex].append(sheet['B' + str(i + 4)].value)
        globalPayouts[tournamentIndex] = list(filter(None, globalPayouts[tournamentIndex]))
        globalPayouts[tournamentIndex].sort() #arranges values in payouts from lowest number to highest
        globalStacks[tournamentIndex] = list(filter(None, globalStacks[tournamentIndex]))

        totalChips = int(sum(globalStacks[tournamentIndex]))
        trimPayouts(tournamentIndex) #cuts out extra payouts that may have been listed in Excel sheet
        numPayouts = len(globalStacks[tournamentIndex])
        payoutList.append({})
        finishDistribution.append({})
        finishProbability.append({})
        globalVarianceList.append({})

        #this only works for tournaments already in the money
        for i in range(1, (1 + numPayouts)):
                globalCDFList.append([])
                ICMValues.append(sheet['C' + str(icmRowNum + i)].value)
                finishDistribution[tournamentIndex].setdefault('Player ' + str(i), {})
                finishProbability[tournamentIndex].setdefault('Player ' + str(i), {})
                globalVarianceList[tournamentIndex].setdefault('Player ' + str(i), {})
                globalVarianceList[tournamentIndex]['Player ' + str(i)].setdefault('Variance')
                globalVarianceList[tournamentIndex]['Player ' + str(i)].setdefault('Percentage of Chips', (globalStacks[tournamentIndex][i - 1] / totalChips))
                globalVarianceList[tournamentIndex]['Player ' + str(i)].setdefault('Standard Deviation')
                globalVarianceList[tournamentIndex]['Player ' + str(i)].setdefault('EV')                
                payoutList[tournamentIndex].setdefault('Player ' + str(i), [])
                for j in range(1, (1 + numPayouts)):
                        #singleICM_Sim uses integer print outs for finDist
                        finishDistribution[tournamentIndex]['Player ' + str(i)].setdefault('Position ' + str(j), int(sheet['G' + str(rowNum + j)].value))
                        #calculation below is going to be slightly off because globalTrials is larger than length of payoutList after read in because of rounding
                        finishProbability[tournamentIndex]['Player ' + str(i)].setdefault('Position ' + str(j), finishDistribution[tournamentIndex]['Player ' + str(i)]['Position ' + str(j)] / globalTrials)
                        payoutList[tournamentIndex]['Player ' + str(i)] += int(finishDistribution[tournamentIndex]['Player ' + str(i)]['Position ' + str(j)]) * [globalPayouts[tournamentIndex][numPayouts - j]]
                rowNum += (numPayouts + 2)
                shuffle(payoutList[tournamentIndex]['Player ' + str(i)])
        return ICMValues

def trimPayouts(tournamentIndex):
        global globalPayouts
        globalPayouts[tournamentIndex] = globalPayouts[tournamentIndex][len(globalPayouts[tournamentIndex]) - len(globalStacks[tournamentIndex]) : len(globalPayouts[tournamentIndex])]
        return

def readExcelInput():
        wb = openpyxl.load_workbook('TournamentData.xlsx')
        sheet = wb['Sheet1']        
        global globalTrials, globalActionPlayer, globalActionPrice, globalActionMarkup, globalSampleRunCount, globalActionPercentage, globalTrialsPerSampleRun, globalNPPList
        global globalEdge, globalBRSize, globalBRSizeMEMS, payoutList, finishDistribution, finishProbability, globalVarianceList, globalCDFList, globalRoRList, doMultiTrialSim, globalOutputSheetNumber, globalCollisionList
        global doVarCalc, doRoRCalc, doLoopRoR, doPlotPMF, doMEMS, readICMValuesFromExcel, globalPayouts, globalUniquePayouts, globalStacks, globalPayoutStepList, icmValues, yearsToSimMEMS, lpInvestmentSizeMEMS, doLPSimMEMS, adjSValueList, truePeriodEVMEMS
        
        readICMValuesFromExcel = sheet['J8'].value 
        globalTrials = sheet['C1'].value
        globalActionPlayer = sheet['J2'].value
        globalActionPrice = 0.00        
        globalActionMarkup = sheet['L2'].value     
        globalActionPercentage = sheet['K2'].value            
        globalSampleRunCount = sheet['L5'].value
        globalTrialsPerSampleRun = sheet['M5'].value
        globalEdge = sheet['M2'].value
        globalBRSize = sheet['N2'].value        
        globalPayouts = []
        globalUniquePayouts = []
        globalPayoutStepList = []
        globalStacks = []
        globalCollisionList = []
        icmValues = []
        payoutList = []
        finishDistribution = []
        finishProbability = []
        globalVarianceList = []
        globalCDFList = [] #this gets built in trial function
        globalRoRList = []
        adjSValueList = []
        globalNPPList = []
        doMultiTrialSim = sheet['K5'].value
        globalOutputSheetNumber = sheet['N5'].value
        doVarCalc = sheet['M8'].value
        doRoRCalc = sheet['L8'].value
        doPlotPMF = sheet['J5'].value
        doLoopRoR = sheet['K8'].value
        truePeriodEVMEMS = 0.00
        doMEMS = sheet['J11'].value
        globalBRSizeMEMS = sheet['K11'].value
        yearsToSimMEMS = sheet['L11'].value
        doLPSimMEMS = sheet['M11'].value
        lpInvestmentSizeMEMS = sheet['N11'].value
        return

def adjS():
        #this function estimates the ICM value of a stack on Day 2 according to the
        #GG Poker AdjS calculation which is: ICM Value = 1.35 * buyin * (Day 2 chips / Day 1 chips)^0.75
        #Need to look into how using the buyin with rake effects this estimation
        startingStack = 25000
        buyin = 1500
        for dayTwoChips in globalStacks[0]:
                adjSValueList.append(1.35 * buyin * (dayTwoChips / startingStack) ** 0.75)
        return

if __name__ == '__main__':
        readExcelInput()
        playerIndex = int(globalActionPlayer[7:12])
        
        if(readICMValuesFromExcel == 'n'):
                tournamentIndex = 0
                icmValues.append([])
                icmValues[tournamentIndex] = sicm('TournamentData.xlsx', tournamentIndex)
                if(globalEdge > 1.00):
                        edgeTransform(globalEdge, playerIndex, tournamentIndex)
        else:
                tournamentIndex = 0
                icmValues.append([])
                icmValues[tournamentIndex] = readICMValues('updatedTournamentData.xlsx', 'Sheet' + str(globalOutputSheetNumber), tournamentIndex)
                #if I want to adjust edge after reading in icmValues, I have to change printDist to print 0EV finishDistribution

        if(globalEdge > 1.00):
                rebuildPayoutList(tournamentIndex)

        globalActionPrice = globalActionMarkup * globalActionPercentage * icmValues[tournamentIndex][playerIndex - 1]
        adjS()
        printDist('TournamentData.xlsx', globalActionPercentage, globalActionMarkup, tournamentIndex)             
        
        if((globalSampleRunCount * globalTrialsPerSampleRun) > globalTrials):
                print("ERROR: Sample run * sample trials exceeds total number of trials simulated")
                sys.exit()
        
        #if(doMultiTrialSim or doVarCalc or doRoRCalc == 'y'):
                #actionResultList, actionProfitList, actionProfitListNormalized = buildActionProfitList(actionPlayer, actionPrice)

        if(doMultiTrialSim == 'y'):
                printActionSim(globalActionPlayer, globalActionPrice, globalActionPercentage, globalSampleRunCount, globalTrialsPerSampleRun)
        if(doVarCalc =='y'):
                varianceCalc(tournamentIndex, playerIndex)
        if(doRoRCalc =='y'):
                calcRiskOfRuin(playerIndex, tournamentIndex)